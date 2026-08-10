from django.shortcuts import render
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from django.db.models import Sum, Avg, Max, Min
from clientes.models import Cliente
from cuentas.models import Cuenta
from movimientos.models import Movimiento
from datetime import datetime
from openpyxl import Workbook
from .utils import encabezado_reporte, pie_pagina
from openpyxl.styles import PatternFill, Alignment

from usuarios.decorators import grupos_permitidos

@grupos_permitidos("Administrador", "Supervisor")
def index(request):
    return render(request, "reportes/index.html")

@grupos_permitidos("Administrador", "Supervisor")
def reporte_dashboard_pdf(request):

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        'attachment; filename="Dashboard_Bancario.pdf"'
    )

    doc = SimpleDocTemplate(response)

    elementos = encabezado_reporte("REPORTE EJECUTIVO DEL BANCO")

    estilos = getSampleStyleSheet()

    elementos.append(

        Paragraph(

            "<b>Últimos movimientos</b>",

            estilos["Heading2"]

        )

    )

    movimientos = (

        Movimiento.objects
        .select_related("cuenta")
        .order_by("-fecha")[:5]

    )

    datos_mov = [

        ["Cuenta", "Tipo", "Monto", "Fecha"]

    ]

    for mov in movimientos:

        datos_mov.append([

            mov.cuenta.numero_cuenta,

            mov.get_tipo_display(),

            f"S/. {mov.monto:,.2f}",

            mov.fecha.strftime("%d/%m/%Y")

        ])

    tabla2 = Table(datos_mov)

    tabla2.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.darkgreen),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.grey),

            ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ])

    )

    elementos.append(tabla2)

    elementos.append(Spacer(1, 20))

    total_clientes = Cliente.objects.count()
    total_cuentas = Cuenta.objects.count()

    saldo_total = (
        Cuenta.objects.aggregate(total=Sum("saldo"))["total"] or 0
    )

    saldo_promedio = (
        Cuenta.objects.aggregate(promedio=Avg("saldo"))["promedio"] or 0
    )

    mayor_saldo = (
        Cuenta.objects.aggregate(maximo=Max("saldo"))["maximo"] or 0
    )

    menor_saldo = (
        Cuenta.objects.aggregate(minimo=Min("saldo"))["minimo"] or 0
    )

    datos = [

        ["Indicador", "Valor"],

        ["Clientes", total_clientes],

        ["Cuentas", total_cuentas],

        ["Saldo total", f"S/. {saldo_total:,.2f}"],

        ["Saldo promedio", f"S/. {saldo_promedio:,.2f}"],

        ["Mayor saldo", f"S/. {mayor_saldo:,.2f}"],

        ["Menor saldo", f"S/. {menor_saldo:,.2f}"],

    ]

    tabla = Table(datos)

    tabla.setStyle(

        TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.darkblue),

            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("GRID", (0,0), (-1,-1), 1, colors.grey),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige),

            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0,0), (-1,0), 10),

        ])

    )

    tabla = Table(datos)

    elementos.append(tabla)

    doc.build(
        elementos,
        onFirstPage=pie_pagina,
        onLaterPages=pie_pagina
    )

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_dashboard_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    ws["A1"] = "REPORTE EJECUTIVO DEL BANCO"

    ws.append([])

    ws.append(["Indicador", "Valor"])

    total_clientes = Cliente.objects.count()
    total_cuentas = Cuenta.objects.count()

    saldo_total = Cuenta.objects.aggregate(
        total=Sum("saldo")
    )["total"] or 0

    saldo_promedio = Cuenta.objects.aggregate(
        promedio=Avg("saldo")
    )["promedio"] or 0

    mayor_saldo = Cuenta.objects.aggregate(
        maximo=Max("saldo")
    )["maximo"] or 0

    menor_saldo = Cuenta.objects.aggregate(
        minimo=Min("saldo")
    )["minimo"] or 0

    ws.append(["Clientes", total_clientes])
    ws.append(["Cuentas", total_cuentas])
    ws.append(["Saldo total", float(saldo_total)])
    ws.append(["Saldo promedio", float(saldo_promedio)])
    ws.append(["Mayor saldo", float(mayor_saldo)])
    ws.append(["Menor saldo", float(menor_saldo)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Dashboard_Bancario.xlsx"'
    )

    wb.save(response)

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_clientes_pdf(request):

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        'attachment; filename="Clientes.pdf"'
    )

    doc = SimpleDocTemplate(response)

    elementos = encabezado_reporte("REPORTE DE CLIENTES")

    datos = [

        [
            "DNI",
            "Nombre",
            "Teléfono",
            "Correo"
        ]

    ]

    clientes = Cliente.objects.all().order_by("apellidos", "nombres")

    for cliente in clientes:

        datos.append([

            cliente.dni,

            f"{cliente.apellidos}, {cliente.nombres}",

            cliente.telefono,

            cliente.correo,

        ])

    tabla = Table(datos)

    tabla.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.darkblue),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.grey),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("BOTTOMPADDING",(0,0),(-1,0),10),

        ])

    )

    tabla = Table(datos)

    elementos.append(tabla)

    doc.build(
        elementos,
        onFirstPage=pie_pagina,
        onLaterPages=pie_pagina
    )

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_clientes_excel(request):

    wb = Workbook()

    ws = wb.active
    ws.title = "Clientes"

    # ============================
    # ENCABEZADOS
    # ============================

    ws.append([
        "N.º",
        "DNI",
        "Nombres",
        "Apellidos",
        "Fecha de nacimiento",
        "Dirección",
        "Teléfono",
        "Correo",
        "Estado",
        "Fecha de registro",
    ])

    # ============================
    # CLIENTES
    # ============================

    clientes = Cliente.objects.all().order_by(
        "apellidos",
        "nombres"
    )

    for numero, cliente in enumerate(clientes, start=1):

        ws.append([

            numero,

            cliente.dni,

            cliente.nombres,

            cliente.apellidos,

            cliente.fecha_nacimiento.strftime(
                "%d/%m/%Y"
            ),

            cliente.direccion,

            cliente.telefono,

            cliente.correo,

            cliente.get_estado_display(),

            cliente.fecha_registro.strftime(
                "%d/%m/%Y %H:%M"
            ),

        ])

    # ============================
    # ESTILO DEL ENCABEZADO
    # ============================

    for celda in ws[1]:

        celda.font = celda.font.copy(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="0B3A75"
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ============================
    # ANCHO DE COLUMNAS
    # ============================

    anchos = {
        "A": 8,
        "B": 12,
        "C": 25,
        "D": 30,
        "E": 22,
        "F": 40,
        "G": 18,
        "H": 35,
        "I": 15,
        "J": 22,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[columna].width = ancho

    # ============================
    # FILTRO
    # ============================

    ws.auto_filter.ref = ws.dimensions

    # ============================
    # CONGELAR ENCABEZADO
    # ============================

    ws.freeze_panes = "A2"

    # ============================
    # RESPUESTA
    # ============================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="Clientes.xlsx"'
    )

    wb.save(response)

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_cuentas_pdf(request):

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="Cuentas.pdf"'
    )

    doc = SimpleDocTemplate(response)

    elementos = encabezado_reporte("REPORTE DE CUENTAS")

    datos = [[
        "N° Cuenta",
        "Cliente",
        "Tipo",
        "Saldo"
    ]]

    cuentas = (
        Cuenta.objects
        .select_related("cliente")
        .order_by("numero_cuenta")
    )

    for cuenta in cuentas:

        datos.append([

            cuenta.numero_cuenta,

            f"{cuenta.cliente.apellidos}, {cuenta.cliente.nombres}",

            cuenta.get_tipo_cuenta_display(),

            f"S/. {cuenta.saldo:,.2f}",

        ])

    tabla = Table(datos)

    tabla.setStyle(

        TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.darkgreen),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

            ("GRID", (0,0), (-1,-1), 1, colors.grey),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige),

            ("BOTTOMPADDING", (0,0), (-1,0), 10),

        ])

    )

    tabla = Table(datos)

    elementos.append(tabla)

    doc.build(
        elementos,
        onFirstPage=pie_pagina,
        onLaterPages=pie_pagina
)

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_movimientos_pdf(request):

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        'attachment; filename="Movimientos.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=40,
    )

    elementos = []

    # ENCABEZADO
    elementos.extend(
        encabezado_reporte(
            "REPORTE DE MOVIMIENTOS BANCARIOS"
        )
    )

    # DATOS DE LA TABLA
    datos = [
        [
            "Cuenta",
            "Cliente",
            "Tipo",
            "Monto",
            "Saldo actual",
            "Fecha",
        ]
    ]

    movimientos = (
        Movimiento.objects
        .select_related(
            "cuenta",
            "cuenta__cliente"
        )
        .order_by("-fecha")
    )

    for movimiento in movimientos:

        cuenta = movimiento.cuenta
        cliente = cuenta.cliente

        datos.append([

            cuenta.numero_cuenta,

            f"{cliente.apellidos}, "
            f"{cliente.nombres}",

            movimiento.get_tipo_display(),

            f"S/. {movimiento.monto:,.2f}",

            f"S/. {movimiento.saldo_actual:,.2f}",

            movimiento.fecha.strftime(
                "%d/%m/%Y %H:%M"
            ),

        ])

    tabla = Table(
        datos,
        repeatRows=1,
        colWidths=[
            65,
            125,
            75,
            70,
            80,
            90,
        ]
    )

    tabla.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkblue
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "BACKGROUND",
                (0, 1),
                (-1, -1),
                colors.beige
            ),

            (
                "ALIGN",
                (3, 1),
                (4, -1),
                "RIGHT"
            ),

            (
                "ALIGN",
                (5, 1),
                (5, -1),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

        ])

    )

    elementos.append(tabla)

    doc.build(
        elementos,
        onFirstPage=pie_pagina,
        onLaterPages=pie_pagina
    )

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_movimientos_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Movimientos"

    # ENCABEZADO
    ws.append([
        "Fecha",
        "Cuenta",
        "Cliente",
        "Tipo",
        "Monto",
        "Saldo anterior",
        "Saldo actual",
        "Descripción",
    ])

    # DATOS
    movimientos = (
        Movimiento.objects
        .select_related(
            "cuenta",
            "cuenta__cliente"
        )
        .order_by("-fecha")
    )

    for movimiento in movimientos:

        cliente = movimiento.cuenta.cliente

        ws.append([

            movimiento.fecha.strftime(
                "%d/%m/%Y %H:%M"
            ),

            movimiento.cuenta.numero_cuenta,

            f"{cliente.apellidos}, "
            f"{cliente.nombres}",

            movimiento.get_tipo_display(),

            float(movimiento.monto),

            float(movimiento.saldo_anterior),

            float(movimiento.saldo_actual),

            movimiento.descripcion,

        ])

    # ESTILO DEL ENCABEZADO

    for celda in ws[1]:

        celda.font = celda.font.copy(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="0B3A75"
        )

        celda.alignment = Alignment(
            horizontal="center"
        )

    # ANCHO DE COLUMNAS

    anchos = {
        "A": 20,
        "B": 18,
        "C": 35,
        "D": 20,
        "E": 15,
        "F": 18,
        "G": 18,
        "H": 40,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[columna].width = ancho

    # FORMATO MONETARIO

    for fila in range(2, ws.max_row + 1):

        ws[f"E{fila}"].number_format = (
            '"S/." #,##0.00'
        )

        ws[f"F{fila}"].number_format = (
            '"S/." #,##0.00'
        )

        ws[f"G{fila}"].number_format = (
            '"S/." #,##0.00'
        )

    # FILTRO

    ws.auto_filter.ref = ws.dimensions

    # CONGELAR ENCABEZADO

    ws.freeze_panes = "A2"

    # RESPUESTA

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="Movimientos.xlsx"'
    )

    wb.save(response)

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_cuentas_excel(request):

    wb = Workbook()

    ws = wb.active
    ws.title = "Cuentas"

    # ============================
    # ENCABEZADOS
    # ============================

    ws.append([
        "N.º",
        "Número de cuenta",
        "Cliente",
        "Tipo",
        "Saldo",
        "Fecha de apertura",
        "Estado",
    ])

    # ============================
    # CUENTAS
    # ============================

    cuentas = (
        Cuenta.objects
        .select_related("cliente")
        .order_by(
            "cliente__apellidos",
            "cliente__nombres"
        )
    )

    for numero, cuenta in enumerate(cuentas, start=1):

        ws.append([

            numero,

            cuenta.numero_cuenta,

            f"{cuenta.cliente.apellidos}, "
            f"{cuenta.cliente.nombres}",

            cuenta.get_tipo_cuenta_display(),

            float(cuenta.saldo),

            cuenta.fecha_apertura.strftime(
                "%d/%m/%Y"
            ),

            cuenta.get_estado_display(),

        ])

    # ============================
    # ESTILO DEL ENCABEZADO
    # ============================

    for celda in ws[1]:

        celda.font = celda.font.copy(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="0B3A75"
        )

        celda.alignment = Alignment(
            horizontal="center"
        )

    # ============================
    # ANCHO DE COLUMNAS
    # ============================

    anchos = {
        "A": 8,
        "B": 20,
        "C": 35,
        "D": 22,
        "E": 18,
        "F": 20,
        "G": 15,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[columna].width = ancho

    # ============================
    # FORMATO MONETARIO
    # ============================

    for fila in range(2, ws.max_row + 1):

        ws[f"E{fila}"].number_format = (
            '"S/." #,##0.00'
        )

    # ============================
    # FILTRO
    # ============================

    ws.auto_filter.ref = ws.dimensions

    # ============================
    # CONGELAR ENCABEZADO
    # ============================

    ws.freeze_panes = "A2"

    # ============================
    # RESPUESTA
    # ============================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="Cuentas.xlsx"'
    )

    wb.save(response)

    return response



