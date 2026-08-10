from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from cuentas.models import Cuenta
from .models import Movimiento
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import os
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.units import cm
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from usuarios.decorators import grupos_permitidos
from auditoria.utils import registrar_auditoria
from django.db.models import Sum

@grupos_permitidos("Administrador", "Cajero")
def deposito(request):

    if request.method == "POST":

        numero_cuenta = request.POST.get("numero_cuenta", "").strip()
        monto_texto = request.POST.get("monto", "").strip()

        if not numero_cuenta:
            messages.error(request, "Debe ingresar el número de cuenta.")
            return redirect("deposito")

        if not monto_texto:
            messages.error(request, "Debe ingresar el monto.")
            return redirect("deposito")

        try:
            monto = Decimal(monto_texto.replace(",", "."))
        except InvalidOperation:
            messages.error(request, "El monto no es válido.")
            return redirect("deposito")
        
        if monto <= 0:
            messages.error(
                request,
                "El monto debe ser mayor que cero."
            )
            return redirect("deposito")

        try:
            with transaction.atomic():

                cuenta = Cuenta.objects.select_for_update().get(
                    numero_cuenta=numero_cuenta
                )
                
                if cuenta.estado != "A":

                    messages.error(
                        request,
                        "La cuenta no está activa."
                    )

                    return redirect("deposito")

                saldo_anterior = cuenta.saldo
                cuenta.saldo += monto
                cuenta.save()

                Movimiento.objects.create(
                    cuenta=cuenta,
                    tipo='D',
                    monto=monto,
                    saldo_anterior=saldo_anterior,
                    saldo_actual=cuenta.saldo,
                    descripcion="Depósito en efectivo"
                )
                
                registrar_auditoria(
                    request,
                    "Movimientos",
                    "Depósito",
                    f"Depósito de S/ {monto:.2f} en cuenta {cuenta.numero_cuenta}"
                )

                messages.success(request, "Depósito realizado correctamente.")
                return redirect("deposito")

        except Cuenta.DoesNotExist:
            messages.error(request, "La cuenta no existe.")

    return render(request, "movimientos/deposito.html")

@grupos_permitidos("Administrador", "Cajero")
def retiro(request):

    if request.method == "POST":

        numero_cuenta = request.POST.get("numero_cuenta")
        
        if not numero_cuenta:

                messages.error(
                    request,
                    "Debe ingresar el número de cuenta."
                )

                return redirect("retiro")
        try:
            monto = Decimal(
                request.POST.get("monto", "").replace(",", ".")
            )
        except InvalidOperation:
            messages.error(request, "Monto inválido")
            return redirect("retiro")
        
        if monto <= 0:
            messages.error(
                request,
                "El monto debe ser mayor que cero."
            )
            return redirect("retiro")

        try:
            with transaction.atomic():

                cuenta = Cuenta.objects.select_for_update().get(
                    numero_cuenta=numero_cuenta
                )
                
                if cuenta.estado != "A":

                    messages.error(
                        request,
                        "La cuenta no está activa."
                    )

                    return redirect("retiro")

                if cuenta.saldo < monto:
                    messages.error(request, "Saldo insuficiente")
                    return redirect("retiro")

                saldo_anterior = cuenta.saldo
                cuenta.saldo -= monto
                cuenta.save()

                Movimiento.objects.create(
                    cuenta=cuenta,
                    tipo='R',
                    monto=monto,
                    saldo_anterior=saldo_anterior,
                    saldo_actual=cuenta.saldo,
                    descripcion="Retiro en efectivo"
                )
                
                registrar_auditoria(
                    request,
                    "Movimientos",
                    "Retiro",
                    f"Retiro de S/ {monto:.2f} de cuenta {cuenta.numero_cuenta}"
                )

                messages.success(request, "Retiro realizado correctamente")
                return redirect("retiro")

        except Cuenta.DoesNotExist:
            messages.error(request, "Cuenta no encontrada")

    return render(request, "movimientos/retiro.html")

@grupos_permitidos("Administrador", "Cajero")
def transferencia(request):

    if request.method == "POST":

        origen_num = request.POST.get("origen")
        destino_num = request.POST.get("destino")
        if origen_num == destino_num:

            messages.error(
                request,
                "La cuenta origen y destino no pueden ser la misma."
            )

            return redirect("transferencia")
        try:

            monto = Decimal(
                request.POST.get("monto", "").replace(",", ".")
            )

        except InvalidOperation:

            messages.error(
                request,
                "Monto inválido"
            )

            return redirect("transferencia")

        if monto <= 0:

            messages.error(
                request,
                "El monto debe ser mayor que cero."
            )

            return redirect("transferencia")

        try:
            with transaction.atomic():

                origen = Cuenta.objects.select_for_update().get(
                    numero_cuenta=origen_num
                )

                destino = Cuenta.objects.select_for_update().get(
                    numero_cuenta=destino_num
                )
                
                if origen.estado != "A":

                    messages.error(
                        request,
                        "La cuenta origen está inactiva."
                    )

                    return redirect("transferencia")

                if destino.estado != "A":

                    messages.error(
                        request,
                        "La cuenta destino está inactiva."
                    )

                    return redirect("transferencia")

                if origen.saldo < monto:
                    messages.error(request, "Saldo insuficiente en cuenta origen")
                    return redirect("transferencia")

                # 🔻 descuento origen
                saldo_origen_anterior = origen.saldo
                origen.saldo -= monto
                origen.save()

                Movimiento.objects.create(
                    cuenta=origen,
                    tipo='T',
                    monto=monto,
                    saldo_anterior=saldo_origen_anterior,
                    saldo_actual=origen.saldo,
                    descripcion=f"Transferencia a {destino.numero_cuenta}"
                )

                # 🔺 abono destino
                saldo_destino_anterior = destino.saldo
                destino.saldo += monto
                destino.save()

                Movimiento.objects.create(
                    cuenta=destino,
                    tipo='T',
                    monto=monto,
                    saldo_anterior=saldo_destino_anterior,
                    saldo_actual=destino.saldo,
                    descripcion=f"Transferencia desde {origen.numero_cuenta}"
                )
                
                registrar_auditoria(
                    request,
                    "Movimientos",
                    "Transferencia",
                    f"Transferencia de S/ {monto:.2f} de {origen.numero_cuenta} a {destino.numero_cuenta}"
                )

                messages.success(request, "Transferencia realizada correctamente")
                return redirect("transferencia")

        except Cuenta.DoesNotExist:
            messages.error(request, "Cuenta origen o destino no existe")

    return render(request, "movimientos/transferencia.html")

def historial(request):

    movimientos = Movimiento.objects.all().order_by('fecha')

    return render(request, "movimientos/historial.html", {
        "movimientos": movimientos
    })
    
@grupos_permitidos("Administrador", "Supervisor")
def reporte_movimientos_pdf(request):
    
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    tipo = request.GET.get("tipo")

    movimientos = Movimiento.objects.select_related(
        "cuenta",
        "cuenta__cliente"
    ).order_by("-fecha")

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        'attachment; filename="reporte_movimientos.pdf"'
    )

    pdf = canvas.Canvas(response, pagesize=A4)

    ANCHO, ALTO = A4

    # ==========================
    # CABECERA
    # ==========================

    logo = os.path.join(
        settings.BASE_DIR,
        "static",
        "img",
        "logo_banco.png"
    )

    pdf.setFillColor(colors.HexColor("#0B3A75"))

    pdf.rect(
        0,
        ALTO - 0.5 * cm,
        ANCHO,
        0.5 * cm,
        fill=1,
        stroke=0
    )

    try:
        pdf.drawImage(
            logo,
            2 * cm,
            ALTO - 3.2 * cm,
            width=2.2 * cm,
            height=2.2 * cm,
            preserveAspectRatio=True,
            mask="auto"
        )
    except:
        pass

    pdf.setFillColor(colors.HexColor("#0B3A75"))
    pdf.setFont("Helvetica-Bold", 22)

    pdf.drawString(
        5 * cm,
        ALTO - 1.8 * cm,
        "BANCO PYTHON"
    )

    pdf.setFont("Helvetica", 11)
    pdf.setFillColor(colors.grey)

    pdf.drawString(
        5 * cm,
        ALTO - 2.5 * cm,
        "Tecnología • Seguridad • Confianza"
    )

    pdf.setFillColor(colors.black)

    pdf.setFont("Helvetica", 9)

    pdf.drawRightString(
        19 * cm,
        ALTO - 1.8 * cm,
        f"Emitido: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    pdf.setStrokeColor(colors.HexColor("#0B3A75"))

    pdf.line(
        2 * cm,
        ALTO - 3.6 * cm,
        19 * cm,
        ALTO - 3.6 * cm
    )

    pdf.setFont("Helvetica-Bold", 18)

    pdf.setFillColor(colors.HexColor("#0B3A75"))

    pdf.drawCentredString(
    ANCHO / 2,
    ALTO - 4.5 * cm,
    "REPORTE GENERAL DE MOVIMIENTOS"
    )

    # ======================================
    # TABLA
    # ======================================

    y = ALTO - 6 * cm

    alto_fila = 0.7 * cm

    x1 = 1.5 * cm   # Cuenta
    x2 = 4.5 * cm   # Cliente
    x3 = 9.0 * cm   # Tipo
    x4 = 12.0 * cm  # Monto
    x5 = 15.0 * cm  # Fecha
    x6 = 18.5 * cm

    pdf.setFillColor(colors.HexColor("#0B3A75"))

    pdf.rect(
        x1,
        y,
        x6 - x1,
        alto_fila,
        fill=1,
        stroke=1
    )

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 9)

    pdf.drawCentredString(3 * cm, y+0.22*cm, "Cuenta")
    pdf.drawCentredString(6.7 * cm, y+0.22*cm, "Cliente")
    pdf.drawCentredString(10.5 * cm, y+0.22*cm, "Tipo")
    pdf.drawCentredString(13.5 * cm, y+0.22*cm, "Monto")
    pdf.drawCentredString(16.7 * cm, y+0.22*cm, "Fecha")

    # ======================================
    # FILAS DE MOVIMIENTOS
    # ======================================

    y -= alto_fila

    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica", 8)

    for mov in movimientos:    
        
        pdf.drawCentredString(
        3 * cm,
        y,
        mov.cuenta.numero_cuenta
        )

        pdf.drawString(
        5.5 * cm,
        y,
        f"{mov.cuenta.cliente.nombres} {mov.cuenta.cliente.apellidos}"[:30]
        )

        pdf.drawString(
            10.2 * cm,
            y,
            mov.get_tipo_display()
        )

        pdf.drawCentredString(
        13.5 * cm,
        y,
        f"S/ {mov.monto:.2f}"
        )

        pdf.drawCentredString(
        16.7 * cm,
        y,
        mov.fecha.strftime("%d/%m/%Y")
        )

        # línea separadora
        pdf.setStrokeColor(colors.lightgrey)

        pdf.line(
            x1,
            y - 0.1 * cm,
            x6,
            y - 0.1 * cm
        )

        y -= alto_fila
        
    # ======================================
    # RESUMEN DE MOVIMIENTOS
    # ======================================

    total_depositos = movimientos.filter(
        tipo="D"
    ).aggregate(total=Sum("monto"))["total"] or 0

    total_retiros = movimientos.filter(
        tipo="R"
    ).aggregate(total=Sum("monto"))["total"] or 0

    total_transferencias = movimientos.filter(
        tipo="T"
    ).aggregate(total=Sum("monto"))["total"] or 0


    y -= 1 * cm

    pdf.setStrokeColor(colors.HexColor("#0B3A75"))

    pdf.line(
        2 * cm,
        y,
        19 * cm,
        y
    )

    y -= 0.8 * cm

    pdf.setFont("Helvetica-Bold", 11)
    pdf.setFillColor(colors.HexColor("#0B3A75"))

    pdf.drawString(
        2 * cm,
        y,
        "RESUMEN DE MOVIMIENTOS"
    )

    y -= 0.6 * cm

    pdf.setFont("Helvetica", 10)
    pdf.setFillColor(colors.black)

    pdf.drawString(
        2 * cm,
        y,
        f"Total depósitos: S/ {total_depositos:,.2f}"
    )

    y -= 0.5 * cm

    pdf.drawString(
        2 * cm,
        y,
        f"Total retiros: S/ {total_retiros:,.2f}"
    )

    y -= 0.5 * cm

    pdf.drawString(
        2 * cm,
        y,
        f"Total transferencias: S/ {total_transferencias:,.2f}"
    )
    
    # ======================================
    # PIE DE PÁGINA
    # ======================================

    pdf.setStrokeColor(colors.HexColor("#0B3A75"))

    pdf.line(
        2 * cm,
        1.5 * cm,
        19 * cm,
        1.5 * cm
    )

    pdf.setFont("Helvetica", 8)

    pdf.setFillColor(colors.grey)

    pdf.drawString(
        2 * cm,
        1 * cm,
        "Banco Python - Reporte generado automáticamente"
    )

    pdf.drawRightString(
        19 * cm,
        1 * cm,
        f"Página 1"
    )

    pdf.save()

    return response

@grupos_permitidos("Administrador", "Supervisor")
def reporte_movimientos_excel(request):

    movimientos = Movimiento.objects.select_related(
        "cuenta",
        "cuenta__cliente"
    ).order_by("-fecha")

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Encabezados
    encabezados = [
        "Fecha",
        "Cuenta",
        "Cliente",
        "Tipo",
        "Monto",
        "Saldo Anterior",
        "Saldo Actual",
        "Descripción"
    ]

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="0B3A75"
        )
        celda.alignment = Alignment(horizontal="center")

    fila = 2

    for m in movimientos:

        ws.cell(row=fila, column=1).value = m.fecha.strftime("%d/%m/%Y %H:%M")
        ws.cell(row=fila, column=2).value = m.cuenta.numero_cuenta
        ws.cell(
            row=fila,
            column=3
        ).value = f"{m.cuenta.cliente.nombres} {m.cuenta.cliente.apellidos}"
        ws.cell(row=fila, column=4).value = m.get_tipo_display()
        ws.cell(row=fila, column=5).value = float(m.monto)
        ws.cell(row=fila, column=6).value = float(m.saldo_anterior)
        ws.cell(row=fila, column=7).value = float(m.saldo_actual)
        ws.cell(row=fila, column=8).value = m.descripcion

        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="movimientos.xlsx"'
    )

    wb.save(response)

    return response
