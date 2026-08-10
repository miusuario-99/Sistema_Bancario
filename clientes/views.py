from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Cliente
from .forms import ClienteForm
from django.core.paginator import Paginator
from django.db.models import Q
from openpyxl import Workbook
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from auditoria.utils import registrar_auditoria
from openpyxl.styles import PatternFill, Alignment
from usuarios.decorators import grupos_permitidos


# ==========================
# LISTAR CLIENTES
# ==========================
@grupos_permitidos("Administrador", "Cajero")
def lista_clientes(request):

    query = request.GET.get("q", "")

    clientes = Cliente.objects.all()

    # 🔍 BUSCADOR
    if query:
        clientes = clientes.filter(
            Q(dni__icontains=query) |
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query)
        )

    # 📄 PAGINACIÓN
    paginator = Paginator(clientes, 5)  # 5 por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ==========================
    # ESTADÍSTICAS
    # ==========================

    total_clientes = Cliente.objects.count()

    clientes_activos = Cliente.objects.filter(
        estado="A"
    ).count()

    clientes_inactivos = Cliente.objects.filter(
        estado="I"
    ).count()

    return render(request, "clientes/lista.html", {

        "page_obj": page_obj,
        "query": query,

        "total_clientes": total_clientes,
        "clientes_activos": clientes_activos,
        "clientes_inactivos": clientes_inactivos,

    })

# ==========================
# NUEVO CLIENTE
# ==========================
@grupos_permitidos("Administrador", "Cajero")
def nuevo_cliente(request):

    if request.method == "POST":

        form = ClienteForm(request.POST)

        if form.is_valid():            
            cliente = form.save()

            registrar_auditoria(
                request,
                "Clientes",
                "Nuevo cliente",
                f"Cliente: {cliente.dni} - {cliente.nombres} {cliente.apellidos}"
            )
            messages.success(request, "Cliente registrado correctamente.")
            return redirect("lista_clientes")

    else:

        form = ClienteForm()

    return render(request, "clientes/formulario.html", {
        "form": form,
        "titulo": "Nuevo Cliente"
    })


# ==========================
# EDITAR CLIENTE
# ==========================
@grupos_permitidos("Administrador", "Cajero")
def editar_cliente(request, id):

    cliente = get_object_or_404(Cliente, id=id)

    if request.method == "POST":

        form = ClienteForm(request.POST, instance=cliente)

        if form.is_valid():
           
            cliente = form.save()

            registrar_auditoria(
                request,
                "Clientes",
                "Editar",
                f"Cliente: {cliente.dni} - {cliente.nombres} {cliente.apellidos}"
            )
            messages.success(request, "Cliente actualizado correctamente.")
            return redirect("lista_clientes")

    else:

        form = ClienteForm(instance=cliente)

    return render(request, "clientes/formulario.html", {
        "form": form,
        "titulo": "Editar"
    })

# ==========================
# ELIMINAR CLIENTE
# ==========================
@grupos_permitidos("Administrador", "Cajero")
def eliminar_cliente(request, id):

    cliente = get_object_or_404(Cliente, id=id)

    cliente.delete()
    registrar_auditoria(
    request,
    "Clientes",
    "Eliminar",
    f"Cliente: {cliente.dni} - {cliente.nombres} {cliente.apellidos}"
    )

    messages.success(
        request,
        "Cliente eliminado correctamente"
    )

    return redirect("lista_clientes")
    
    # ==========================
    # EXPORTAR CLIENTES A EXCEL
    # ==========================

#@grupos_permitidos("Administrador", "Cajero")
def exportar_clientes_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Clientes"

    # ============================
    # ENCABEZADOS
    # ============================

    ws.append([
        "N.º",
        "DNI",
        "Nombres",
        "Apellidos",
        "Fecha de nacimiento",
        "Dirección",
        "Teléfono",
        "Correo",
        "Estado",
        "Fecha de registro",
    ])

    # ============================
    # CLIENTES
    # ============================

    clientes = Cliente.objects.all().order_by(
        "apellidos",
        "nombres"
    )

    for numero, cliente in enumerate(clientes, start=1):

        estado = (
            "Activo"
            if cliente.estado == "A"
            else "Inactivo"
        )

        ws.append([

            numero,

            cliente.dni,

            cliente.nombres,

            cliente.apellidos,

            cliente.fecha_nacimiento.strftime(
                "%d/%m/%Y"
            ),

            cliente.direccion,

            cliente.telefono,

            cliente.correo,

            estado,

            cliente.fecha_registro.strftime(
                "%d/%m/%Y %H:%M"
            ),

        ])

    # ============================
    # ESTILO DEL ENCABEZADO
    # ============================

    for celda in ws[1]:

        celda.font = celda.font.copy(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="0B3A75"
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ============================
    # ANCHO DE COLUMNAS
    # ============================

    anchos = {
        "A": 8,
        "B": 12,
        "C": 25,
        "D": 30,
        "E": 22,
        "F": 40,
        "G": 18,
        "H": 35,
        "I": 15,
        "J": 22,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[columna].width = ancho

    # ============================
    # FILTRO
    # ============================

    ws.auto_filter.ref = ws.dimensions

    # ============================
    # CONGELAR ENCABEZADO
    # ============================

    ws.freeze_panes = "A2"

    # ============================
    # RESPUESTA
    # ============================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="Clientes.xlsx"'
    )

    wb.save(response)

    return response

# ==========================
# EXPORTAR CLIENTES PDF
# ==========================

#@grupos_permitidos("Administrador", "Cajero")
def exportar_clientes_pdf(request):

    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        Image
    )

    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from django.conf import settings
    from datetime import datetime
    import os


    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="clientes.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=letter
    )

    elementos = []

    estilos = getSampleStyleSheet()

    # ==========================
    # LOGO
    # ==========================

    ruta_logo = os.path.join(
        settings.BASE_DIR,
        "static",
        "img",
        "logo_banco.png"
    )

    if os.path.exists(ruta_logo):

        logo = Image(
            ruta_logo,
            width=90,
            height=90
        )

        elementos.append(logo)

    elementos.append(
        Spacer(1, 10)
    )


    # ==========================
    # TITULO
    # ==========================

    elementos.append(

        Paragraph(
            "Banco Python",
            estilos["Title"]
        )

    )


    elementos.append(

        Paragraph(
            "Reporte General de Clientes",
            estilos["Heading2"]
        )

    )


    fecha = datetime.now().strftime(
        "%d/%m/%Y %H:%M"
    )


    elementos.append(

        Paragraph(
            f"Fecha de emisión: {fecha}",
            estilos["Normal"]
        )

    )


    elementos.append(
        Spacer(1,20)
    )


    # ==========================
    # TABLA
    # ==========================

    datos = [

        [
            "N°",
            "DNI",
            "Nombres",
            "Apellidos",
            "Teléfono",
            "Estado"
        ]

    ]


    clientes = Cliente.objects.all().order_by(
        "apellidos"
    )


    contador = 1


    for cliente in clientes:

        datos.append([

            contador,
            cliente.dni,
            cliente.nombres,
            cliente.apellidos,
            cliente.telefono,
            "Activo" if cliente.estado == "A"
            else "Inactivo"

        ])

        contador += 1

    tabla = Table(
        datos,
        repeatRows=1
    )

    tabla.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.darkblue
            ),

            (
                "TEXTCOLOR",
                (0,0),
                (-1,0),
                colors.white
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0,0),
                (-1,-1),
                "CENTER"
            )

        ])

    )


    elementos.append(tabla)

    # ==========================
    # PIE
    # ==========================

    elementos.append(
        Spacer(1,30)
    )

    elementos.append(

        Paragraph(
            "Banco Python © 2026 - Sistema Bancario Administrativo",
            estilos["Normal"]
        )

    )

    doc.build(elementos)

    return response
