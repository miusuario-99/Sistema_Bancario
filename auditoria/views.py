from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Auditoria
from django.http import HttpResponse

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image
)

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

from django.conf import settings
from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from usuarios.decorators import grupos_permitidos

@grupos_permitidos("Administrador", "Supervisor")
def lista_auditoria(request):

    query = request.GET.get("q", "")

    auditorias = Auditoria.objects.select_related(
        "usuario"
    ).order_by("-fecha")

    if query:

        auditorias = auditorias.filter(

            Q(usuario__username__icontains=query) |
            Q(modulo__icontains=query) |
            Q(accion__icontains=query) |
            Q(descripcion__icontains=query)

        )

    paginator = Paginator(auditorias, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "auditoria/lista.html",
        {
            "page_obj": page_obj,
            "query": query,
        }
    )
    
# ==========================
# EXPORTAR AUDITORÍA PDF
# ==========================

@grupos_permitidos("Administrador", "Supervisor")
def exportar_auditoria_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="auditoria.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=40,
    )

    elementos = []

    estilos = getSampleStyleSheet()

    # ============================
    # LOGO
    # ============================

    ruta_logo = os.path.join(
        settings.BASE_DIR,
        "static",
        "img",
        "logo_banco.png"
    )

    if os.path.exists(ruta_logo):

        logo = Image(
            ruta_logo,
            width=70,
            height=70
        )

        elementos.append(logo)

    elementos.append(
        Spacer(1, 10)
    )

    # ============================
    # ENCABEZADO
    # ============================

    elementos.append(
        Paragraph(
            "<b>BANCO PYTHON</b>",
            estilos["Title"]
        )
    )

    elementos.append(
        Paragraph(
            "Sistema de Gestión Bancaria",
            estilos["Heading2"]
        )
    )

    elementos.append(
        Paragraph(
            "<b>REPORTE DE AUDITORÍA</b>",
            estilos["Heading2"]
        )
    )

    elementos.append(
        Paragraph(
            f"Fecha de emisión: "
            f"{datetime.now().strftime('%d/%m/%Y %H:%M')}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Spacer(1, 20)
    )

    # ============================
    # DATOS
    # ============================

    datos = [[
        "N.º",
        "Usuario",
        "Módulo",
        "Acción",
        "Descripción",
        "Fecha",
    ]]

    auditorias = (
        Auditoria.objects
        .select_related("usuario")
        .order_by("-fecha")
    )

    for numero, registro in enumerate(
        auditorias,
        start=1
    ):

        usuario = (
            registro.usuario.username
            if registro.usuario
            else "Sistema"
        )

        datos.append([

            str(numero),

            usuario,

            registro.modulo,

            registro.accion,

            registro.descripcion or "",

            registro.fecha.strftime(
                "%d/%m/%Y %H:%M"
            ),

        ])

    # ============================
    # TABLA
    # ============================

    tabla = Table(
        datos,
        repeatRows=1,
        colWidths=[
            30,
            70,
            65,
            85,
            160,
            85,
        ]
    )

    tabla.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkblue
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "BACKGROUND",
                (0, 1),
                (-1, -1),
                colors.whitesmoke
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "ALIGN",
                (0, 0),
                (0, -1),
                "CENTER"
            ),

            (
                "ALIGN",
                (5, 1),
                (5, -1),
                "CENTER"
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

        ])
    )

    elementos.append(tabla)

    elementos.append(
        Spacer(1, 20)
    )

    elementos.append(
        Paragraph(
            "Banco Python © 2026 - "
            "Sistema Bancario Administrativo",
            estilos["Normal"]
        )
    )

    doc.build(elementos)

    return response

# ==========================
# EXPORTAR AUDITORÍA EXCEL
# ==========================

@grupos_permitidos("Administrador", "Supervisor")
def exportar_auditoria_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Auditoría"

    # ============================
    # ENCABEZADOS
    # ============================

    ws.append([
        "N.º",
        "Usuario",
        "Módulo",
        "Acción",
        "Descripción",
        "Fecha",
    ])

    # ============================
    # AUDITORÍA
    # ============================

    auditorias = (
        Auditoria.objects
        .select_related("usuario")
        .order_by("-fecha")
    )

    for numero, registro in enumerate(
        auditorias,
        start=1
    ):

        usuario = (
            registro.usuario.username
            if registro.usuario
            else "Sistema"
        )

        ws.append([

            numero,

            usuario,

            registro.modulo,

            registro.accion,

            registro.descripcion or "",

            registro.fecha.strftime(
                "%d/%m/%Y %H:%M"
            ),

        ])

    # ============================
    # ESTILO ENCABEZADO
    # ============================

    for celda in ws[1]:

        celda.font = celda.font.copy(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="0B3A75"
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ============================
    # ANCHO DE COLUMNAS
    # ============================

    anchos = {
        "A": 8,
        "B": 20,
        "C": 20,
        "D": 30,
        "E": 50,
        "F": 22,
    }

    for columna, ancho in anchos.items():

        ws.column_dimensions[
            columna
        ].width = ancho

    # ============================
    # FILTRO
    # ============================

    ws.auto_filter.ref = ws.dimensions

    # ============================
    # CONGELAR ENCABEZADO
    # ============================

    ws.freeze_panes = "A2"

    # ============================
    # RESPUESTA
    # ============================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="Auditoria.xlsx"'
    )

    wb.save(response)

    return response
